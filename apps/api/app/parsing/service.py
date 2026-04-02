from __future__ import annotations

import csv
from datetime import date, datetime, time
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from app.parsing.schemas import ParsedUploadResult
from app.uploads.schemas import FileFormat, SourceColumnPreview, Upload

_SAMPLE_ROW_LIMIT = 5


class UploadParseError(Exception):
    """Raised when an uploaded source file cannot be parsed into RoadViz preview rows."""


def _cell_to_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        normalized = value.strip()
        return normalized or None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return str(value)


def _is_empty_row(values: Iterable[str | None]) -> bool:
    return all(value in {None, ""} for value in values)


def _infer_column_type(values: list[str | None]) -> str:
    populated_values = [value for value in values if value not in {None, ""}]
    if not populated_values:
        return "text"

    for value in populated_values:
        try:
            float(value)
        except (TypeError, ValueError):
            return "text"

    return "number"


def _normalize_headers(raw_headers: list[str | None]) -> list[str]:
    headers: list[str] = []
    seen_headers: dict[str, str] = {}

    for index, header in enumerate(raw_headers, start=1):
        normalized = (header or "").strip() or f"column_{index}"
        dedupe_key = normalized.casefold()
        if dedupe_key in seen_headers:
            raise UploadParseError(
                f"Duplicate source column header '{normalized}' was found in the uploaded file."
            )
        seen_headers[dedupe_key] = normalized
        headers.append(normalized)

    return headers


def _coerce_row_to_record(headers: list[str], raw_values: list[str | None], row_number: int) -> dict[str, str | None]:
    if len(raw_values) > len(headers) and any(value not in {None, ""} for value in raw_values[len(headers) :]):
        raise UploadParseError(
            f"Row {row_number} has more populated values than the header row defines."
        )

    values = raw_values[: len(headers)]
    if len(values) < len(headers):
        values.extend([None] * (len(headers) - len(values)))

    return {
        header: value
        for header, value in zip(headers, values, strict=True)
    }


class UploadParsingService:
    """Read stored upload files and expose real preview rows for mapping and normalization."""

    def parse_upload(self, upload: Upload, storage_path: str | None) -> ParsedUploadResult:
        if upload.file_format == FileFormat.UNKNOWN:
            raise UploadParseError(
                "Unsupported file format. RoadViz preview currently supports CSV and XLSX uploads."
            )
        if not storage_path:
            raise UploadParseError("No stored file is available for this upload.")

        file_path = Path(storage_path)
        if not file_path.exists():
            raise UploadParseError("The uploaded source file could not be found on disk.")

        rows = self._read_rows(upload.file_format, file_path)
        if not rows:
            raise UploadParseError("The uploaded file is empty or does not contain a header row.")

        header_index = next(
            (index for index, row in enumerate(rows) if not _is_empty_row(row)),
            None,
        )
        if header_index is None:
            raise UploadParseError("The uploaded file does not contain any readable rows.")

        headers = _normalize_headers(rows[header_index])
        parsed_rows: list[dict[str, str | None]] = []
        source_samples: dict[str, list[str | None]] = {header: [] for header in headers}

        for offset, raw_row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            record = _coerce_row_to_record(headers, raw_row, offset)
            if _is_empty_row(record.values()):
                continue
            parsed_rows.append(record)
            for header in headers:
                source_samples[header].append(record.get(header))

        source_columns = [
            SourceColumnPreview(
                name=header,
                sample_values=source_samples[header][:3],
                inferred_type=_infer_column_type(source_samples[header]),
            )
            for header in headers
        ]

        return ParsedUploadResult(
            upload_id=upload.id,
            filename=upload.filename,
            file_format=upload.file_format,
            source_columns=source_columns,
            sample_rows=parsed_rows[:_SAMPLE_ROW_LIMIT],
            row_count=len(parsed_rows),
            rows=parsed_rows,
        )

    def _read_rows(self, file_format: FileFormat, file_path: Path) -> list[list[str | None]]:
        if file_format == FileFormat.CSV:
            return self._read_csv(file_path)
        if file_format == FileFormat.XLSX:
            return self._read_xlsx(file_path)
        raise UploadParseError(
            "Unsupported file format. RoadViz preview currently supports CSV and XLSX uploads."
        )

    def _read_csv(self, file_path: Path) -> list[list[str | None]]:
        try:
            with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.reader(handle)
                return [[_cell_to_text(value) for value in row] for row in reader]
        except (OSError, UnicodeDecodeError, csv.Error) as exc:
            raise UploadParseError(f"CSV parsing failed: {exc}") from exc

    def _read_xlsx(self, file_path: Path) -> list[list[str | None]]:
        try:
            workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        except Exception as exc:  # pragma: no cover - openpyxl raises several parse exceptions
            raise UploadParseError(f"XLSX parsing failed: {exc}") from exc

        try:
            worksheet = workbook[workbook.sheetnames[0]]
            return [
                [_cell_to_text(value) for value in row]
                for row in worksheet.iter_rows(values_only=True)
            ]
        finally:
            workbook.close()
